import os
import re
import requests
import urllib.parse
from urllib.parse import urljoin
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# ====================== 配置区（根据实际情况调整）======================
# 注意：如果在本地运行，替换为你的Excel实际路径；当前适配上传的文件路径
EXCEL_PATH = r"D:\脚本需要的文件\江西预算网站.xlsx"  # 上传文件的路径（云端）
# 本地运行时替换为：EXCEL_PATH = r"D:\脚本需要的文件\江西预算网站.xlsx"
SAVE_DIR = r"D:\多网站脚本下载\预算\江西部门"  # 本地保存路径
# ==================================================================

def get_url_from_hyperlink(cell):
    """从Excel单元格超链接中提取真实网址"""
    try:
        if cell.hyperlink is not None:
            return cell.hyperlink.target
    except Exception as e:
        print(f"提取超链接失败：{e}")
        pass
    return None

def fix_chinese_filename(filename):
    """修复中文文件名乱码（处理URL编码、UTF-8解码问题）"""
    try:
        # 情况1：如果文件名是URL编码格式（如%E4%B8%AD%E5%9B%BD），先解码
        decoded_filename = urllib.parse.unquote(filename)
        # 情况2：如果是UTF-8编码被错误解析为GBK，尝试转码修复
        if any(ord(c) > 127 for c in decoded_filename):
            return decoded_filename
        else:
            # 尝试GBK转UTF-8修复
            return decoded_filename.encode('latin-1').decode('utf-8')
    except:
        # 修复失败时返回原始文件名（清理非法字符）
        invalid_chars = r'[\\/:*?"<>|]'
        return re.sub(invalid_chars, "", filename)

def download_from_url(site_name, target_url):
    """下载单个网站的PDF（含中文文件名修复）"""
    print(f"\n========================================")
    print(f"正在处理网站：{site_name}")
    print(f"网址：{target_url}")
    print(f"========================================\n")

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
        "Accept-Language": "zh-CN,zh;q=0.9"  # 告诉服务器优先返回中文内容
    }

    # 访问网页
    try:
        resp = requests.get(target_url, headers=headers, timeout=20)
        resp.raise_for_status()
        # 修复网页内容编码（避免解析时中文乱码）
        if 'charset' not in resp.encoding.lower():
            resp.encoding = resp.apparent_encoding
    except Exception as e:
        print(f"❌ 访问失败：{e}")
        return

    # 解析PDF链接
    soup = BeautifulSoup(resp.text, "html.parser")

    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        # 修复链接文本的中文编码
        link_text = a.get_text(strip=True)
        link_text = fix_chinese_filename(link_text)

        # 只处理PDF链接
        if not href.lower().endswith(".pdf"):
            continue

        # 跳过含“汇总”的PDF
        if "汇总" in href or "汇总" in link_text:
            print(f"⏭️ 跳过汇总文件：{link_text}")
            continue

        # 拼接完整PDF地址
        full_url = urljoin(target_url, href)

        # 生成最终文件名（中文正常显示，避免重复后缀）
        filename = link_text + ".pdf"
        # 避免出现双重.pdf后缀（如“文件.pdf.pdf”）
        filename = re.sub(r'\.pdf\.pdf$', '.pdf', filename)
        # 再次清理非法字符
        invalid_chars = r'[\\/:*?"<>|]'
        filename = re.sub(invalid_chars, "", filename)
        save_path = os.path.join(SAVE_DIR, filename)

        # 跳过已下载的文件
        if os.path.exists(save_path):
            print(f"✅ 已存在，跳过：{filename}")
            continue

        # 下载PDF文件
        print(f"📥 下载中：{filename}")
        try:
            pdf_resp = requests.get(full_url, headers=headers, stream=True, timeout=30)
            pdf_resp.raise_for_status()
            with open(save_path, "wb") as f:
                for chunk in pdf_resp.iter_content(chunk_size=1024*1024):
                    if chunk:
                        f.write(chunk)
            print(f"✅ 下载完成：{filename}\n")
        except Exception as e:
            print(f"❌ 下载失败：{filename} | 错误：{e}\n")

def read_excel_and_download():
    """读取Excel并批量下载PDF"""
    # 创建保存目录（本地路径）
    if not os.path.exists(SAVE_DIR):
        os.makedirs(SAVE_DIR, exist_ok=True)
        print(f"✅ 创建保存目录：{SAVE_DIR}")

    # 读取Excel文件
    try:
        wb = load_workbook(EXCEL_PATH, data_only=True)
        ws = wb.active
        print(f"✅ 成功读取Excel：{EXCEL_PATH}")
    except Exception as e:
        print(f"❌ Excel打开失败：{e}")
        return

    # 遍历A列所有行（提取网站名称和超链接）
    row_count = 0
    for row in ws.iter_rows(min_row=1, max_col=1, values_only=False):
        cell = row[0]
        # 获取网站名称（修复中文编码）
        site_name = str(cell.value).strip() if cell.value else f"未命名网站_{row_count+1}"
        site_name = fix_chinese_filename(site_name)
        # 获取超链接
        url = get_url_from_hyperlink(cell)

        # 跳过无URL的行
        if not url or not isinstance(url, str):
            print(f"⏭️ 跳过无效行（无URL）：{site_name}")
            row_count += 1
            continue

        # 处理URL（补全http/https）
        if not url.startswith(("http://", "https://")):
            url = "http://" + url  # 默认补http（避免相对路径）

        # 开始下载该网站的PDF
        download_from_url(site_name, url)
        row_count += 1

    print(f"\n🎉 所有网站处理完毕！共处理 {row_count} 行数据")

if __name__ == "__main__":
    read_excel_and_download()