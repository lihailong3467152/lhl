import os
import re
import requests
from urllib.parse import urljoin
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# ====================== 固定配置 ======================
EXCEL_PATH = r"D:\脚本需要的文件\任意网站.xlsx"  # 你的Excel路径
SAVE_DIR = r"D:\多网站脚本下载\预算\下载"                # 保存PDF的路径
# ======================================================

def get_url_from_hyperlink(cell):
    """从Excel单元格超链接中提取真实网址"""
    try:
        if cell.hyperlink is not None:
            return cell.hyperlink.target
    except:
        pass
    return None

def download_from_url(site_name, target_url):
    """下载单个网站的PDF"""
    print(f"\n========================================")
    print(f"正在处理网站：{site_name}")
    print(f"网址：{target_url}")
    print(f"========================================\n")

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"
    }

    # 访问网页
    try:
        resp = requests.get(target_url, headers=headers, timeout=20)
        resp.raise_for_status()
    except Exception as e:
        print(f"❌ 访问失败：{e}")
        return

    # 解析PDF
    soup = BeautifulSoup(resp.text, "html.parser")

    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        link_text = a.get_text(strip=True)

        # 只处理PDF
        if not href.lower().endswith(".pdf"):
            continue

        # 跳过含“汇总”
        if "汇总" in href or "汇总" in link_text:
            print(f"⏭️ 跳过汇总：{link_text}")
            continue

        # 拼接完整地址
        full_url = urljoin(target_url, href)

        # 文件名：网页显示的标题 + .pdf
        filename = link_text + ".pdf"
        # 清理非法字符
        invalid_chars = r'[\\/:*?"<>|]'
        filename = re.sub(invalid_chars, "", filename)
        save_path = os.path.join(SAVE_DIR, filename)

        # 跳过已下载
        if os.path.exists(save_path):
            print(f"✅ 已存在：{filename}")
            continue

        # 下载
        print(f"📥 下载中：{filename}")
        try:
            pdf_resp = requests.get(full_url, headers=headers, stream=True, timeout=30)
            with open(save_path, "wb") as f:
                f.write(pdf_resp.content)
            print(f"✅ 下载完成：{filename}\n")
        except Exception as e:
            print(f"❌ 下载失败：{filename} | {e}\n")

def read_excel_and_download():
    """读取Excel并批量下载"""
    # 创建保存目录
    if not os.path.exists(SAVE_DIR):
        os.makedirs(SAVE_DIR)

    # 读取Excel
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
    except Exception as e:
        print(f"Excel打开失败：{e}")
        return

    # 遍历每一行（A列读取）
    for row in ws.iter_rows(min_row=1, max_col=1):
        cell = row[0]
        site_name = str(cell.value).strip() if cell.value else ""
        url = get_url_from_hyperlink(cell)

        if not url:
            continue

        # 开始处理
        download_from_url(site_name, url)

    print("\n🎉 所有网站处理完毕！")

if __name__ == "__main__":
    read_excel_and_download()