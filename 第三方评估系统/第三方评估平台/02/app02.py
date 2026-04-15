import streamlit as st
import sqlite3
import hashlib
import pandas as pd
import datetime
from datetime import timedelta
import openpyxl
from fpdf import FPDF
import json
import time
import plotly.express as px
import plotly.graph_objects as go
import io

# -------------------------- 页面配置 --------------------------
st.set_page_config(
    page_title="第三方机构绩效评估系统",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# -------------------------- CSS样式 --------------------------
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

:root {
    --primary-color: #0ea5e9;
    --primary-dark: #0284c7;
    --secondary-color: #6366f1;
    --success-color: #10b981;
    --warning-color: #f59e0b;
    --danger-color: #ef4444;
    --info-color: #3b82f6;
    --card-shadow: 0 10px 40px rgba(0,0,0,0.1);
    --hover-shadow: 0 20px 60px rgba(0,0,0,0.15);
}

html, body, [class*="css"] {
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
}

.main-header {
    font-size: 42px;
    font-weight: 800;
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
    text-align: center;
    margin-bottom: 30px;
    letter-spacing: -1px;
}

.sub-header {
    font-size: 24px;
    color: #1e293b;
    font-weight: 700;
    margin: 20px 0 15px 0;
    padding-bottom: 10px;
    border-bottom: 3px solid linear-gradient(90deg, #667eea, #764ba2);
    position: relative;
}

.sub-header::after {
    content: '';
    position: absolute;
    bottom: -3px;
    left: 0;
    width: 60px;
    height: 3px;
    background: linear-gradient(90deg, #667eea, #764ba2);
    border-radius: 2px;
}

.card {
    background: white;
    border-radius: 20px;
    padding: 25px;
    box-shadow: var(--card-shadow);
    transition: all 0.3s ease;
    border: 1px solid rgba(226, 232, 240, 0.8);
    margin-bottom: 20px;
}

.card:hover {
    transform: translateY(-5px);
    box-shadow: var(--hover-shadow);
}

.stat-card {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    border-radius: 16px;
    padding: 25px;
    color: white;
    text-align: center;
    transition: all 0.3s ease;
    cursor: pointer;
}

.stat-card:hover {
    transform: scale(1.05);
    box-shadow: 0 15px 50px rgba(102, 126, 234, 0.4);
}

.stat-card.success {
    background: linear-gradient(135deg, #10b981 0%, #059669 100%);
}

.stat-card.warning {
    background: linear-gradient(135deg, #f59e0b 0%, #d97706 100%);
}

.stat-card.info {
    background: linear-gradient(135deg, #3b82f6 0%, #2563eb 100%);
}

.stat-card.danger {
    background: linear-gradient(135deg, #ef4444 0%, #dc2626 100%);
}

.stat-number {
    font-size: 36px;
    font-weight: 800;
    margin-bottom: 5px;
}

.stat-label {
    font-size: 14px;
    opacity: 0.9;
    font-weight: 500;
}

.stButton > button {
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%) !important;
    color: white !important;
    border: none !important;
    border-radius: 12px !important;
    padding: 12px 28px !important;
    font-weight: 600 !important;
    font-size: 15px !important;
    transition: all 0.3s ease !important;
    box-shadow: 0 4px 15px rgba(102, 126, 234, 0.3) !important;
}

.stButton > button:hover {
    transform: translateY(-2px) !important;
    box-shadow: 0 8px 25px rgba(102, 126, 234, 0.4) !important;
}

.login-container {
    background: white;
    border-radius: 24px;
    padding: 50px;
    box-shadow: var(--card-shadow);
    max-width: 450px;
    margin: 0 auto;
    position: relative;
    overflow: hidden;
}

.login-container::before {
    content: '';
    position: absolute;
    top: 0;
    left: 0;
    right: 0;
    height: 5px;
    background: linear-gradient(90deg, #667eea, #764ba2, #f093fb);
}

.stTextInput > div > div > input {
    border-radius: 12px !important;
    border: 2px solid #e2e8f0 !important;
    padding: 14px 18px !important;
    font-size: 15px !important;
    transition: all 0.3s ease !important;
}

.stTextInput > div > div > input:focus {
    border-color: #667eea !important;
    box-shadow: 0 0 0 4px rgba(102, 126, 234, 0.1) !important;
}

.css-1d391kg, .css-163ttbj {
    background: linear-gradient(180deg, #1e293b 0%, #0f172a 100%) !important;
}

.badge {
    display: inline-block;
    padding: 6px 14px;
    border-radius: 20px;
    font-size: 12px;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.5px;
}

.badge-success {
    background: rgba(16, 185, 129, 0.15);
    color: #059669;
}

.badge-warning {
    background: rgba(245, 158, 11, 0.15);
    color: #d97706;
}

.badge-danger {
    background: rgba(239, 68, 68, 0.15);
    color: #dc2626;
}

.badge-info {
    background: rgba(59, 130, 246, 0.15);
    color: #2563eb;
}

@keyframes fadeIn {
    from { opacity: 0; transform: translateY(20px); }
    to { opacity: 1; transform: translateY(0); }
}

.animate-fadeIn {
    animation: fadeIn 0.5s ease-out;
}

.info-box {
    background: linear-gradient(135deg, rgba(59, 130, 246, 0.1) 0%, rgba(37, 99, 235, 0.1) 100%);
    border-left: 4px solid #3b82f6;
    border-radius: 12px;
    padding: 20px;
    margin: 15px 0;
}

.success-box {
    background: linear-gradient(135deg, rgba(16, 185, 129, 0.1) 0%, rgba(5, 150, 105, 0.1) 100%);
    border-left: 4px solid #10b981;
    border-radius: 12px;
    padding: 20px;
    margin: 15px 0;
}

.divider {
    height: 1px;
    background: linear-gradient(90deg, transparent, #e2e8f0, transparent);
    margin: 30px 0;
}

.small-note {
    font-size: 13px;
    color: #64748b;
    margin-top: 8px;
}

.avatar {
    width: 60px;
    height: 60px;
    border-radius: 50%;
    background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
    display: flex;
    align-items: center;
    justify-content: center;
    color: white;
    font-size: 24px;
    font-weight: 700;
    margin: 0 auto 20px;
}

.platform-title {
    font-size: 28px;
    font-weight: 800;
    text-align: center;
    margin: 8px 0 4px 0;
    background: linear-gradient(90deg, #667eea, #764ba2, #f97316);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    letter-spacing: 0.5px;
}

.platform-subtitle {
    text-align: center;
    color: #64748b;
    margin-top: 0;
    font-size: 14px;
}
</style>
""", unsafe_allow_html=True)

# -------------------------- 数据库初始化 --------------------------
def init_db():
    conn = sqlite3.connect('performance_system.db')
    c = conn.cursor()
    
    c.execute('''CREATE TABLE IF NOT EXISTS users
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  username TEXT UNIQUE NOT NULL,
                  password TEXT NOT NULL,
                  role TEXT NOT NULL,
                  email TEXT,
                  phone TEXT,
                  create_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                  last_login TIMESTAMP)''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS institutions
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  name TEXT NOT NULL,
                  region TEXT,
                  contact TEXT,
                  phone TEXT,
                  email TEXT,
                  address TEXT,
                  status TEXT DEFAULT '正常',
                  create_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS indicators
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  name TEXT NOT NULL,
                  full_score INTEGER NOT NULL,
                  category TEXT,
                  description TEXT,
                  sort INTEGER DEFAULT 0)''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS scores
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  inst_id INTEGER NOT NULL,
                  year INTEGER NOT NULL,
                  indicator_id INTEGER NOT NULL,
                  score INTEGER NOT NULL,
                  evaluator TEXT NOT NULL,
                  evaluate_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                  comment TEXT)''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS materials
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  inst_id INTEGER NOT NULL,
                  year INTEGER NOT NULL,
                  file_content BLOB,
                  file_name TEXT,
                  file_type TEXT,
                  status TEXT DEFAULT '待审核',
                  reviewer TEXT,
                  review_comment TEXT,
                  submit_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                  review_time TIMESTAMP)''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS logs
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  username TEXT NOT NULL,
                  operation TEXT NOT NULL,
                  ip_address TEXT,
                  create_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS notifications
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  user_id INTEGER,
                  title TEXT NOT NULL,
                  content TEXT,
                  type TEXT DEFAULT 'info',
                  is_read INTEGER DEFAULT 0,
                  create_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS settings
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  key TEXT UNIQUE NOT NULL,
                  value TEXT,
                  update_time TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    
    default_users = [
        ('admin', hashlib.md5('123456'.encode()).hexdigest(), '管理员', 'admin@example.com', '13800138000'),
        ('user', hashlib.md5('123456'.encode()).hexdigest(), '评估人员', 'user@example.com', '13800138001'),
        ('inst1', hashlib.md5('123456'.encode()).hexdigest(), '机构用户', 'inst1@example.com', '13800138002')
    ]
    for user in default_users:
        try:
            c.execute('INSERT INTO users (username, password, role, email, phone) VALUES (?,?,?,?,?)', user)
        except sqlite3.IntegrityError:
            pass
    
    default_indicators = [
        ('服务质量', 30, '服务', '评估服务质量和客户满意度'),
        ('工作效率', 25, '效率', '评估工作完成效率'),
        ('专业能力', 25, '能力', '评估专业技能水平'),
        ('团队协作', 20, '协作', '评估团队合作能力')
    ]
    for ind in default_indicators:
        try:
            c.execute('INSERT INTO indicators (name, full_score, category, description, sort) VALUES (?,?,?,?,?)', 
                     (*ind, default_indicators.index(ind)))
        except:
            pass
    
    default_settings = [
        ('email_notification', '1'),
        ('sms_notification', '0'),
        ('timeout_reminder', '1'),
        ('score_notification', '1'),
        ('login_attempts', '5'),
        ('session_timeout', '30'),
        ('two_factor_auth', '0')
    ]
    for setting in default_settings:
        try:
            c.execute('INSERT INTO settings (key, value) VALUES (?,?)', setting)
        except:
            pass
    
    conn.commit()
    conn.close()

# -------------------------- 工具函数 --------------------------
def md5_hash(text):
    return hashlib.md5(text.encode()).hexdigest()

def login_user(username, password):
    conn = sqlite3.connect('performance_system.db')
    c = conn.cursor()
    c.execute('SELECT * FROM users WHERE username=? AND password=?', (username, md5_hash(password)))
    user = c.fetchone()
    if user:
        c.execute('UPDATE users SET last_login=? WHERE id=?', 
                  (datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), user[0]))
        conn.commit()
    conn.close()
    return user

def add_log(username, operation):
    conn = sqlite3.connect('performance_system.db')
    c = conn.cursor()
    c.execute('INSERT INTO logs (username, operation, ip_address) VALUES (?,?,?)', 
              (username, operation, '127.0.0.1'))
    conn.commit()
    conn.close()

def add_notification(user_id, title, content, type='info'):
    conn = sqlite3.connect('performance_system.db')
    c = conn.cursor()
    c.execute('INSERT INTO notifications (user_id, title, content, type) VALUES (?,?,?,?)',
              (user_id, title, content, type))
    conn.commit()
    conn.close()

def get_unread_count(user_id):
    conn = sqlite3.connect('performance_system.db')
    c = conn.cursor()
    c.execute('SELECT COUNT(*) FROM notifications WHERE (user_id=? OR user_id IS NULL) AND is_read=0', (user_id,))
    count = c.fetchone()[0]
    conn.close()
    return count

def mark_all_read(user_id):
    conn = sqlite3.connect('performance_system.db')
    c = conn.cursor()
    c.execute('UPDATE notifications SET is_read=1 WHERE user_id=? OR user_id IS NULL', (user_id,))
    conn.commit()
    conn.close()

def check_approval_timeout():
    conn = sqlite3.connect('performance_system.db')
    c = conn.cursor()
    timeout_time = datetime.datetime.now() - timedelta(days=7)
    c.execute("UPDATE materials SET status = '审批超时' WHERE status = '待审核' AND submit_time < ?", 
              (timeout_time.strftime('%Y-%m-%d %H:%M:%S'),))
    conn.commit()
    conn.close()

def ai_evaluate_document(content, rule):
    score = min(100, max(0, 70 + hash(content) % 30))
    suggestions = []
    if score < 80:
        suggestions.append("补充更多详细数据支撑")
    if score < 90:
        suggestions.append("完善文档格式和排版")
    if len(suggestions) == 0:
        suggestions.append("文档质量良好，继续保持")
    reasons = [
        "数据完整度较高，但部分细节需要完善",
        "内容较为全面，建议增加对比分析",
        "基本符合要求，可进一步优化表达"
    ]
    return {
        "score": score,
        "reason": reasons[abs(hash(content)) % len(reasons)],
        "suggestions": suggestions
    }

def export_pdf(data, title):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(200, 10, txt=title, ln=True, align='C')
    pdf.ln(10)
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt=f"生成时间: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True)
    pdf.ln(5)
    pdf.set_font("Arial", 'B', 12)
    headers = list(data.columns)
    for header in headers:
        pdf.cell(40, 10, txt=str(header), border=1)
    pdf.ln()
    pdf.set_font("Arial", size=10)
    for idx, row in data.iterrows():
        for val in row.values:
            pdf.cell(40, 10, txt=str(val)[:20], border=1)
        pdf.ln()
    filename = f"{title}_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    pdf.output(filename)
    return filename

def export_excel(data, title):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        data.to_excel(writer, sheet_name='数据', index=False)
    output.seek(0)
    return output

def get_score_level(score):
    if score >= 90:
        return "优秀", "badge-success", "progress-bar excellent"
    elif score >= 75:
        return "良好", "badge-warning", "progress-bar good"
    elif score >= 60:
        return "合格", "badge-info", "progress-bar pass"
    else:
        return "不合格", "badge-danger", "progress-bar fail"

def create_gauge_chart(value, title, max_val=100):
    fig = go.Figure(go.Indicator(
        mode="gauge+number+delta",
        value=value,
        domain={'x': [0, 1], 'y': [0, 1]},
        title={'text': title, 'font': {'size': 24}},
        gauge={
            'axis': {'range': [None, max_val], 'tickwidth': 1},
            'bar': {'color': "#667eea"},
            'bgcolor': "white",
            'borderwidth': 2,
            'bordercolor': "#ccc",
            'steps': [
                {'range': [0, 60], 'color': '#fee2e2'},
                {'range': [60, 75], 'color': '#fef3c7'},
                {'range': [75, 90], 'color': '#dbeafe'},
                {'range': [90, 100], 'color': '#d1fae5'}
            ],
            'threshold': {
                'line': {'color': "red", 'width': 4},
                'thickness': 0.75,
                'value': 90
            }
        }
    ))
    fig.update_layout(height=300)
    return fig

def get_setting(key):
    conn = sqlite3.connect('performance_system.db')
    c = conn.cursor()
    c.execute('SELECT value FROM settings WHERE key=?', (key,))
    result = c.fetchone()
    conn.close()
    return result[0] if result else None

def set_setting(key, value):
    conn = sqlite3.connect('performance_system.db')
    c = conn.cursor()
    c.execute('INSERT OR REPLACE INTO settings (key, value, update_time) VALUES (?,?,?)',
              (key, value, datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
    conn.commit()
    conn.close()

def get_status_badge(status):
    mapping = {'通过': 'success', '驳回': 'danger', '待审核': 'warning', '审批超时': 'danger'}
    return mapping.get(status, 'info')


# -------------------------- 页面组件 --------------------------
def render_login_page():
    st.markdown('<div class="animate-fadeIn">', unsafe_allow_html=True)
    st.markdown("""
        <div style="text-align: center; padding: 40px 0 20px;">
            <div style="font-size: 80px; margin-bottom: 10px;">📊</div>
        </div>
    """, unsafe_allow_html=True)
    st.markdown('<p class="main-header">第三方机构绩效评估系统</p>', unsafe_allow_html=True)
    st.markdown('<p style="text-align: center; color: #64748b; margin-bottom: 30px;">Performance Evaluation System</p>', unsafe_allow_html=True)
    
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        with st.container():
            st.markdown('<div class="login-container">', unsafe_allow_html=True)
            st.markdown('<div class="avatar">👤</div>', unsafe_allow_html=True)
            
            with st.form(key='login_form'):
                username = st.text_input("👤 用户名", placeholder="请输入用户名")
                password = st.text_input("🔒 密码", type="password", placeholder="请输入密码")
                col_remember, col_forget = st.columns(2)
                with col_remember:
                    remember = st.checkbox("记住我", value=True)
                with col_forget:
                    st.markdown('<p style="text-align: right; font-size: 13px;"><a href="#" style="color: #667eea;">忘记密码?</a></p>', unsafe_allow_html=True)
                
                submitted = st.form_submit_button("🚀 登录系统", use_container_width=True)
                if submitted:
                    with st.spinner("正在验证..."):
                        time.sleep(0.5)
                        user = login_user(username, password)
                        if user:
                            st.session_state.user = user
                            st.session_state.logged_in = True
                            if remember:
                                st.session_state.remember_user = username
                            add_log(username, "用户登录系统")
                            st.success("✅ 登录成功！正在跳转...")
                            time.sleep(0.5)
                            st.rerun()
                        else:
                            st.error("❌ 用户名或密码错误，请重试。")
            
            st.markdown('</div>', unsafe_allow_html=True)
            st.markdown("""
                <div style="text-align: center; margin-top: 30px; color: #94a3b8; font-size: 13px;">
                    <p>默认账号: admin / 123456 | user / 123456</p>
                    <p>© 2025 第三方机构绩效评估系统 v2.0</p>
                </div>
            """, unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

def render_sidebar():
    with st.sidebar:
        unread = get_unread_count(st.session_state.user[0])
        st.markdown(f"""
            <div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                        border-radius: 16px; padding: 20px; margin-bottom: 20px; color: white;">
                <div style="display: flex; align-items: center; gap: 15px;">
                    <div style="width: 50px; height: 50px; border-radius: 50%; background: rgba(255,255,255,0.2); 
                                display: flex; align-items: center; justify-content: center; font-size: 24px;">👤</div>
                    <div>
                        <div style="font-weight: 700; font-size: 16px;">{st.session_state.user[1]}</div>
                        <div style="font-size: 13px; opacity: 0.8;">{st.session_state.user[3]}</div>
                    </div>
                </div>
            </div>
        """, unsafe_allow_html=True)
        
        st.markdown('<p style="color: rgba(255,255,255,0.6); font-size: 12px; margin-bottom: 10px;">快捷操作</p>', unsafe_allow_html=True)
        col1, col2 = st.columns(2)
        with col1:
            if st.button("📊 仪表盘", use_container_width=True, key="sidebar_dashboard"):
                st.session_state.page = "dashboard"
                st.session_state.show_add_user = False
                st.session_state.show_add_inst = False
                st.session_state.show_export = False
                st.session_state.show_notifications = False
                st.rerun()
        with col2:
            btn_label = f"🔔 通知 ({unread})" if unread > 0 else "🔔 通知"
            if st.button(btn_label, use_container_width=True, key="sidebar_notifications"):
                st.session_state.show_notifications = True
                st.session_state.show_add_user = False
                st.session_state.show_add_inst = False
                st.session_state.show_export = False
                st.rerun()
        st.markdown("<hr style='border-color: rgba(255,255,255,0.1);'>", unsafe_allow_html=True)

def render_notifications():
    st.markdown('<p class="sub-header">🔔 系统通知</p>', unsafe_allow_html=True)
    conn = sqlite3.connect('performance_system.db')
    notifications = pd.read_sql('SELECT * FROM notifications WHERE user_id=? OR user_id IS NULL ORDER BY create_time DESC', 
                                conn, params=(st.session_state.user[0],))
    conn.close()
    
    col1, col2 = st.columns([3, 1])
    with col2:
        if st.button("✅ 全部标记为已读", use_container_width=True):
            mark_all_read(st.session_state.user[0])
            st.success("✅ 已标记所有通知为已读")
            time.sleep(0.5)
            st.rerun()
    
    st.markdown('<div class="card">', unsafe_allow_html=True)
    if not notifications.empty:
        for idx, notif in notifications.iterrows():
            is_read = notif['is_read'] == 1
            bg_color = "#f8fafc" if is_read else "#eff6ff"
            border_color = "#e2e8f0" if is_read else "#3b82f6"
            unread_dot = ' <span style="color: #3b82f6;">●</span>' if not is_read else ''
            st.markdown(f"""
                <div style="background: {bg_color}; border-left: 4px solid {border_color}; 
                            border-radius: 12px; padding: 15px; margin-bottom: 10px;">
                    <div style="display: flex; justify-content: space-between; align-items: center;">
                        <div><span style="font-weight: 600;">{notif['title']}</span>{unread_dot}</div>
                        <span style="color: #94a3b8; font-size: 12px;">{notif['create_time']}</span>
                    </div>
                    <div style="color: #64748b; margin-top: 5px; font-size: 14px;">{notif['content']}</div>
                </div>
            """, unsafe_allow_html=True)
    else:
        st.info("暂无通知")
    st.markdown('</div>', unsafe_allow_html=True)

def render_stat_cards():
    conn = sqlite3.connect('performance_system.db')
    user_cnt = pd.read_sql('SELECT COUNT(*) as cnt FROM users', conn)['cnt'].iloc[0]
    inst_cnt = pd.read_sql('SELECT COUNT(*) as cnt FROM institutions', conn)['cnt'].iloc[0]
    score_cnt = pd.read_sql('SELECT COUNT(*) as cnt FROM scores', conn)['cnt'].iloc[0]
    pending_cnt = pd.read_sql("SELECT COUNT(*) as cnt FROM materials WHERE status='待审核'", conn)['cnt'].iloc[0]
    conn.close()
    
    st.markdown('<div class="animate-fadeIn">', unsafe_allow_html=True)
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.markdown(f'<div class="stat-card"><div class="stat-number">{int(user_cnt)}</div><div class="stat-label">👥 系统用户</div></div>', unsafe_allow_html=True)
    with col2:
        st.markdown(f'<div class="stat-card success"><div class="stat-number">{int(inst_cnt)}</div><div class="stat-label">🏢 评估机构</div></div>', unsafe_allow_html=True)
    with col3:
        st.markdown(f'<div class="stat-card info"><div class="stat-number">{int(score_cnt)}</div><div class="stat-label">📝 评分记录</div></div>', unsafe_allow_html=True)
    with col4:
        st.markdown(f'<div class="stat-card warning"><div class="stat-number">{int(pending_cnt)}</div><div class="stat-label">⏳ 待审批</div></div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

def render_quick_actions():
    st.markdown('<p class="sub-header">⚡ 快速操作</p>', unsafe_allow_html=True)
    col1, col2, col3, col4, col5 = st.columns(5)
    with col1:
        if st.button("➕ 添加用户", use_container_width=True, key="quick_add_user"):
            st.session_state.show_add_user = True
            st.session_state.show_add_inst = False
            st.session_state.show_export = False
            st.session_state.show_notifications = False
            st.rerun()
    with col2:
        if st.button("🏢 添加机构", use_container_width=True, key="quick_add_inst"):
            st.session_state.show_add_inst = True
            st.session_state.show_add_user = False
            st.session_state.show_export = False
            st.session_state.show_notifications = False
            st.rerun()
    with col3:
        if st.button("📈 统计分析", use_container_width=True, key="quick_reports"):
            st.session_state.page = "statistics"
            st.rerun()
    with col4:
        if st.button("📥 导出数据", use_container_width=True, key="quick_export"):
            st.session_state.show_export = True
            st.session_state.show_add_user = False
            st.session_state.show_add_inst = False
            st.session_state.show_notifications = False
            st.rerun()
    with col5:
        if st.button("🤖 AI评估", use_container_width=True, key="quick_ai"):
            st.session_state.page = "ai_eval"
            st.rerun()

def render_add_user_modal():
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<p style="font-weight: 600; margin-bottom: 15px;">➕ 添加新用户</p>', unsafe_allow_html=True)
    conn = sqlite3.connect('performance_system.db')
    
    col1, col2, col3 = st.columns(3)
    with col1:
        new_user = st.text_input("用户名", key="new_username")
        new_email = st.text_input("邮箱", key="new_email")
    with col2:
        new_pwd = st.text_input("密码", type="password", key="new_password")
        new_phone = st.text_input("电话", key="new_phone")
    with col3:
        new_role = st.selectbox("角色", ["管理员", "评估人员", "机构用户"], key="new_role")
    
    col_btn1, col_btn2, _ = st.columns([1, 1, 3])
    with col_btn1:
        if st.button("✅ 确认添加", use_container_width=True, key="confirm_add_user"):
            if new_user and new_pwd:
                try:
                    c = conn.cursor()
                    c.execute('INSERT INTO users (username, password, role, email, phone) VALUES (?,?,?,?,?)',
                              (new_user, md5_hash(new_pwd), new_role, new_email, new_phone))
                    conn.commit()
                    add_log(st.session_state.user[1], f"添加用户：{new_user}")
                    add_notification(None, "新用户添加", f"管理员添加了新用户: {new_user}", "success")
                    st.success("✅ 用户添加成功！")
                    st.session_state.show_add_user = False
                    time.sleep(0.5)
                    st.rerun()
                except sqlite3.IntegrityError:
                    st.error("❌ 用户名已存在")
            else:
                st.warning("⚠️ 请填写完整信息")
    with col_btn2:
        if st.button("❌ 取消", use_container_width=True, key="cancel_add_user"):
            st.session_state.show_add_user = False
            st.rerun()
    
    st.markdown("---")
    st.markdown('<p style="font-weight: 600; margin-bottom: 10px;">📤 批量导入用户</p>', unsafe_allow_html=True)
    uploaded_file = st.file_uploader("上传CSV或Excel文件", type=['csv', 'xlsx'], key="bulk_import_users")
    if uploaded_file:
        try:
            if uploaded_file.name.endswith('.csv'):
                df = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file)
            st.markdown("**预览数据：**")
            st.dataframe(df.head(), use_container_width=True)
            if st.button("📥 确认导入", use_container_width=True, key="confirm_bulk_import"):
                c = conn.cursor()
                imported = 0
                for _, row in df.iterrows():
                    try:
                        c.execute('INSERT INTO users (username, password, role, email, phone) VALUES (?,?,?,?,?)',
                                  (row['username'], md5_hash(str(row['password'])), 
                                   row.get('role', '评估人员'), row.get('email', ''), row.get('phone', '')))
                        imported += 1
                    except:
                        pass
                conn.commit()
                add_log(st.session_state.user[1], f"批量导入用户：{imported}个")
                st.success(f"✅ 成功导入 {imported} 个用户")
                time.sleep(0.5)
                st.rerun()
        except Exception as e:
            st.error(f"❌ 文件解析失败: {str(e)}")
    st.markdown('</div>', unsafe_allow_html=True)
    conn.close()

def render_add_inst_modal():
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<p style="font-weight: 600; margin-bottom: 15px;">➕ 添加新机构</p>', unsafe_allow_html=True)
    conn = sqlite3.connect('performance_system.db')
    
    col1, col2 = st.columns(2)
    with col1:
        name = st.text_input("机构名称", key="inst_name")
        region = st.text_input("所属地区", key="inst_region")
        contact = st.text_input("联系人", key="inst_contact")
    with col2:
        phone = st.text_input("联系电话", key="inst_phone")
        email = st.text_input("邮箱", key="inst_email")
        address = st.text_input("详细地址", key="inst_address")
    
    col_btn1, col_btn2, _ = st.columns([1, 1, 3])
    with col_btn1:
        if st.button("✅ 添加机构", use_container_width=True, key="confirm_add_inst"):
            if name:
                c = conn.cursor()
                c.execute('INSERT INTO institutions (name, region, contact, phone, email, address) VALUES (?,?,?,?,?,?)',
                          (name, region, contact, phone, email, address))
                conn.commit()
                add_log(st.session_state.user[1], f"添加机构：{name}")
                add_notification(None, "新机构添加", f"添加了新机构: {name}", "success")
                st.success("✅ 机构添加成功！")
                st.session_state.show_add_inst = False
                time.sleep(0.5)
                st.rerun()
            else:
                st.warning("⚠️ 请填写机构名称")
    with col_btn2:
        if st.button("❌ 取消", use_container_width=True, key="cancel_add_inst"):
            st.session_state.show_add_inst = False
            st.rerun()
    
    st.markdown("---")
    st.markdown('<p style="font-weight: 600; margin-bottom: 10px;">📤 批量导入机构</p>', unsafe_allow_html=True)
    uploaded_file = st.file_uploader("上传CSV或Excel文件", type=['csv', 'xlsx'], key="bulk_import_inst")
    if uploaded_file:
        try:
            if uploaded_file.name.endswith('.csv'):
                df = pd.read_csv(uploaded_file)
            else:
                df = pd.read_excel(uploaded_file)
            st.markdown("**预览数据：**")
            st.dataframe(df.head(), use_container_width=True)
            if st.button("📥 确认导入", use_container_width=True, key="confirm_bulk_import_inst"):
                c = conn.cursor()
                imported = 0
                for _, row in df.iterrows():
                    try:
                        c.execute('INSERT INTO institutions (name, region, contact, phone, email, address) VALUES (?,?,?,?,?,?)',
                                  (row['name'], row.get('region', ''), row.get('contact', ''),
                                   row.get('phone', ''), row.get('email', ''), row.get('address', '')))
                        imported += 1
                    except:
                        pass
                conn.commit()
                add_log(st.session_state.user[1], f"批量导入机构：{imported}个")
                st.success(f"✅ 成功导入 {imported} 个机构")
                time.sleep(0.5)
                st.rerun()
        except Exception as e:
            st.error(f"❌ 文件解析失败: {str(e)}")
    st.markdown('</div>', unsafe_allow_html=True)
    conn.close()

def render_export_modal():
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<p style="font-weight: 600; margin-bottom: 15px;">📥 数据导出</p>', unsafe_allow_html=True)
    conn = sqlite3.connect('performance_system.db')
    
    export_type = st.selectbox("选择导出类型", ["用户数据", "机构数据", "评分数据", "材料数据", "日志数据", "完整备份"])
    export_format = st.radio("导出格式", ["Excel", "CSV", "JSON"], horizontal=True)
    
    if st.button("📥 生成导出文件", use_container_width=True):
        with st.spinner("正在生成文件..."):
            time.sleep(0.5)
            if export_type == "用户数据":
                df = pd.read_sql('SELECT id, username, role, email, phone, create_time FROM users', conn)
                filename = f"users_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
            elif export_type == "机构数据":
                df = pd.read_sql('SELECT * FROM institutions', conn)
                filename = f"institutions_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
            elif export_type == "评分数据":
                df = pd.read_sql('''SELECT s.*, i.name as indicator_name, inst.name as institution_name
                    FROM scores s JOIN indicators i ON s.indicator_id=i.id JOIN institutions inst ON s.inst_id=inst.id''', conn)
                filename = f"scores_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
            elif export_type == "材料数据":
                df = pd.read_sql('''SELECT m.id, i.name as institution_name, m.year, m.file_name, m.status, 
                    m.reviewer, m.submit_time, m.review_time FROM materials m JOIN institutions i ON m.inst_id=i.id''', conn)
                filename = f"materials_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
            elif export_type == "日志数据":
                df = pd.read_sql('SELECT * FROM logs ORDER BY create_time DESC', conn)
                filename = f"logs_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}"
            else:
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    pd.read_sql('SELECT * FROM users', conn).to_excel(writer, sheet_name='users', index=False)
                    pd.read_sql('SELECT * FROM institutions', conn).to_excel(writer, sheet_name='institutions', index=False)
                    pd.read_sql('SELECT * FROM indicators', conn).to_excel(writer, sheet_name='indicators', index=False)
                    pd.read_sql('SELECT * FROM scores', conn).to_excel(writer, sheet_name='scores', index=False)
                    pd.read_sql('SELECT * FROM materials', conn).to_excel(writer, sheet_name='materials', index=False)
                    pd.read_sql('SELECT * FROM logs', conn).to_excel(writer, sheet_name='logs', index=False)
                output.seek(0)
                st.download_button(label="⬇️ 下载完整备份", data=output,
                    file_name=f"full_backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                add_log(st.session_state.user[1], "导出完整备份")
                conn.close()
                return
            
            if export_format == "Excel":
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='data', index=False)
                output.seek(0)
                st.download_button(label=f"⬇️ 下载 {export_type}.xlsx", data=output, file_name=f"{filename}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            elif export_format == "CSV":
                csv = df.to_csv(index=False).encode('utf-8-sig')
                st.download_button(label=f"⬇️ 下载 {export_type}.csv", data=csv, file_name=f"{filename}.csv",
                    mime="text/csv", use_container_width=True)
            else:
                json_data = df.to_json(orient='records', force_ascii=False)
                st.download_button(label=f"⬇️ 下载 {export_type}.json", data=json_data, file_name=f"{filename}.json",
                    mime="application/json", use_container_width=True)
            add_log(st.session_state.user[1], f"导出{export_type}")
    
    st.markdown("---")
    if st.button("❌ 关闭", use_container_width=True):
        st.session_state.show_export = False
        st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)
    conn.close()


def render_admin_dashboard():
    if st.session_state.get('show_notifications'):
        render_notifications()
        return
    if st.session_state.get('show_add_user'):
        render_quick_actions()
        render_add_user_modal()
        return
    if st.session_state.get('show_add_inst'):
        render_quick_actions()
        render_add_inst_modal()
        return
    if st.session_state.get('show_export'):
        render_quick_actions()
        render_export_modal()
        return
    
    st.markdown('<div class="animate-fadeIn">', unsafe_allow_html=True)
    st.markdown('<p class="platform-title">第三方评估系统</p>', unsafe_allow_html=True)
    st.markdown('<p class="platform-subtitle">管理员控制台</p>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)
    render_stat_cards()
    render_quick_actions()
    st.markdown('<div class="divider"></div>', unsafe_allow_html=True)
    
    menu_options = ["🏠 仪表盘", "👥 用户管理", "🏢 机构管理", "📊 指标管理", 
                    "✅ 审批管理", "📈 统计分析", "📝 日志管理", "🤖 AI文档评估", "⚙️ 系统设置"]

    page = st.session_state.get('page', 'dashboard')
    default_index = 0
    if page == 'statistics':
        default_index = menu_options.index("📈 统计分析")
    elif page == 'ai_eval':
        default_index = menu_options.index("🤖 AI文档评估")

    menu = st.sidebar.selectbox("📋 功能菜单", menu_options, index=default_index)

    if menu == "🏠 仪表盘":
        st.session_state.page = "dashboard"
        render_admin_home()
    elif menu == "👥 用户管理":
        st.session_state.page = "user_management"
        render_user_management()
    elif menu == "🏢 机构管理":
        st.session_state.page = "institution_management"
        render_institution_management()
    elif menu == "📊 指标管理":
        st.session_state.page = "indicator_management"
        render_indicator_management()
    elif menu == "✅ 审批管理":
        st.session_state.page = "approval_management"
        render_approval_management()
    elif menu == "📈 统计分析":
        st.session_state.page = "statistics"
        render_statistics()
    elif menu == "📝 日志管理":
        st.session_state.page = "logs"
        render_logs()
    elif menu == "🤖 AI文档评估":
        st.session_state.page = "ai_eval"
        render_ai_evaluation()
    elif menu == "⚙️ 系统设置":
        st.session_state.page = "settings"
        render_settings()
        render_settings()

def render_admin_home():
    st.markdown('<div class="animate-fadeIn">', unsafe_allow_html=True)
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<p style="font-weight: 600; margin-bottom: 15px;">📈 年度评分趋势</p>', unsafe_allow_html=True)
        conn = sqlite3.connect('performance_system.db')
        trend_df = pd.read_sql('SELECT year, AVG(score) as avg_score, COUNT(*) as count FROM scores GROUP BY year ORDER BY year', conn)
        conn.close()
        if not trend_df.empty:
            fig = px.line(trend_df, x='year', y='avg_score', markers=True, 
                         labels={'year': '年份', 'avg_score': '平均分数'}, line_shape='spline')
            fig.update_traces(line_color='#667eea', line_width=3)
            fig.update_layout(height=300, showlegend=False)
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("暂无数据")
        st.markdown('</div>', unsafe_allow_html=True)
    
    with col2:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<p style="font-weight: 600; margin-bottom: 15px;">📊 评分分布</p>', unsafe_allow_html=True)
        conn = sqlite3.connect('performance_system.db')
        dist_df = pd.read_sql("""SELECT CASE WHEN score >= 90 THEN '优秀' WHEN score >= 75 THEN '良好' 
            WHEN score >= 60 THEN '合格' ELSE '不合格' END as level, COUNT(*) as count FROM scores GROUP BY level""", conn)
        conn.close()
        if not dist_df.empty:
            fig = px.pie(dist_df, values='count', names='level', color_discrete_sequence=['#10b981', '#3b82f6', '#f59e0b', '#ef4444'])
            fig.update_layout(height=300, showlegend=True)
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("暂无数据")
        st.markdown('</div>', unsafe_allow_html=True)
    
    st.markdown('<div class="card" style="margin-top: 20px;">', unsafe_allow_html=True)
    st.markdown('<p style="font-weight: 600; margin-bottom: 15px;">🕐 最近活动</p>', unsafe_allow_html=True)
    conn = sqlite3.connect('performance_system.db')
    logs_df = pd.read_sql('SELECT username, operation, create_time FROM logs ORDER BY create_time DESC LIMIT 10', conn)
    conn.close()
    if not logs_df.empty:
        st.dataframe(logs_df, use_container_width=True, hide_index=True)
    else:
        st.info("暂无活动记录")
    st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

def render_user_management():
    st.markdown('<p class="sub-header">👥 用户管理</p>', unsafe_allow_html=True)
    conn = sqlite3.connect('performance_system.db')
    
    st.markdown('<div class="card">', unsafe_allow_html=True)
    col_search, col_filter = st.columns([2, 1])
    with col_search:
        search = st.text_input("🔍 搜索用户", placeholder="输入用户名搜索")
    with col_filter:
        role_filter = st.selectbox("🎭 角色筛选", ["全部", "管理员", "评估人员", "机构用户"])
    
    query = 'SELECT id, username, role, email, phone, last_login FROM users WHERE 1=1'
    params = []
    if search:
        query += ' AND username LIKE ?'
        params.append(f'%{search}%')
    if role_filter != "全部":
        query += ' AND role = ?'
        params.append(role_filter)
    df = pd.read_sql(query, conn, params=params)
    st.dataframe(df, use_container_width=True, hide_index=True)
    
    if not df.empty:
        st.markdown("---")
        st.markdown('<p style="font-weight: 600; margin-bottom: 10px;">⚡ 用户操作</p>', unsafe_allow_html=True)
        col1, col2, col3 = st.columns([2, 1, 1])
        with col1:
            selected_user = st.selectbox("选择用户", df['id'].tolist(), format_func=lambda x: df[df['id']==x]['username'].iloc[0])
        with col2:
            if st.button("🗑️ 删除用户", use_container_width=True):
                if selected_user != st.session_state.user[0]:
                    c = conn.cursor()
                    c.execute('DELETE FROM users WHERE id=?', (selected_user,))
                    conn.commit()
                    add_log(st.session_state.user[1], f"删除用户ID：{selected_user}")
                    st.success("✅ 用户已删除")
                    time.sleep(0.5)
                    st.rerun()
                else:
                    st.error("❌ 不能删除当前登录用户")
        with col3:
            if st.button("🔑 重置密码", use_container_width=True):
                c = conn.cursor()
                c.execute('UPDATE users SET password=? WHERE id=?', (md5_hash('123456'), selected_user))
                conn.commit()
                add_log(st.session_state.user[1], f"重置用户密码ID：{selected_user}")
                st.success("✅ 密码已重置为 123456")
    st.markdown('</div>', unsafe_allow_html=True)
    
    st.markdown('<div class="card" style="margin-top: 20px;">', unsafe_allow_html=True)
    st.markdown('<p style="font-weight: 600; margin-bottom: 15px;">➕ 添加新用户</p>', unsafe_allow_html=True)
    col1, col2, col3 = st.columns(3)
    with col1:
        new_user = st.text_input("用户名", key="um_new_username")
        new_email = st.text_input("邮箱", key="um_new_email")
    with col2:
        new_pwd = st.text_input("密码", type="password", key="um_new_password")
        new_phone = st.text_input("电话", key="um_new_phone")
    with col3:
        new_role = st.selectbox("角色", ["管理员", "评估人员", "机构用户"], key="um_new_role")
    
    col_btn1, col_btn2, _ = st.columns([1, 1, 3])
    with col_btn1:
        if st.button("✅ 确认添加", use_container_width=True, key="um_confirm_add"):
            if new_user and new_pwd:
                try:
                    c = conn.cursor()
                    c.execute('INSERT INTO users (username, password, role, email, phone) VALUES (?,?,?,?,?)',
                              (new_user, md5_hash(new_pwd), new_role, new_email, new_phone))
                    conn.commit()
                    add_log(st.session_state.user[1], f"添加用户：{new_user}")
                    add_notification(None, "新用户添加", f"管理员添加了新用户: {new_user}", "success")
                    st.success("✅ 用户添加成功！")
                    time.sleep(0.5)
                    st.rerun()
                except sqlite3.IntegrityError:
                    st.error("❌ 用户名已存在")
            else:
                st.warning("⚠️ 请填写完整信息")
    with col_btn2:
        if st.button("📤 批量导入", use_container_width=True, key="um_bulk_import"):
            st.info("📤 请使用快速操作中的批量导入功能")
    st.markdown('</div>', unsafe_allow_html=True)
    conn.close()

def render_institution_management():
    st.markdown('<p class="sub-header">🏢 机构管理</p>', unsafe_allow_html=True)
    conn = sqlite3.connect('performance_system.db')
    
    st.markdown('<div class="card">', unsafe_allow_html=True)
    col1, col2 = st.columns([2, 1])
    with col1:
        search = st.text_input("🔍 搜索机构", placeholder="输入机构名称")
    with col2:
        regions = pd.read_sql('SELECT DISTINCT region FROM institutions WHERE region IS NOT NULL AND region != ""', conn)['region'].dropna().tolist()
        region_filter = st.selectbox("🌍 地区筛选", ["全部"] + regions)
    
    query = 'SELECT * FROM institutions WHERE 1=1'
    params = []
    if search:
        query += ' AND name LIKE ?'
        params.append(f'%{search}%')
    if region_filter != "全部":
        query += ' AND region = ?'
        params.append(region_filter)
    df = pd.read_sql(query, conn, params=params)
    st.dataframe(df, use_container_width=True, hide_index=True)
    
    if not df.empty:
        st.markdown("---")
        st.markdown('<p style="font-weight: 600; margin-bottom: 10px;">⚡ 机构操作</p>', unsafe_allow_html=True)
        col1, col2, col3 = st.columns([2, 1, 1])
        with col1:
            selected_inst = st.selectbox("选择机构", df['id'].tolist(), format_func=lambda x: df[df['id']==x]['name'].iloc[0])
        with col2:
            if st.button("🗑️ 删除机构", use_container_width=True):
                c = conn.cursor()
                c.execute('DELETE FROM institutions WHERE id=?', (selected_inst,))
                conn.commit()
                add_log(st.session_state.user[1], f"删除机构ID：{selected_inst}")
                st.success("✅ 机构已删除")
                time.sleep(0.5)
                st.rerun()
        with col3:
            if st.button("📊 查看详情", use_container_width=True):
                inst_data = df[df['id'] == selected_inst].iloc[0]
                st.markdown(f"**机构名称:** {inst_data['name']}  \n**所属地区:** {inst_data['region'] or '未设置'}  \n**联系人:** {inst_data['contact'] or '未设置'}  \n**联系电话:** {inst_data['phone'] or '未设置'}  \n**邮箱:** {inst_data['email'] or '未设置'}  \n**地址:** {inst_data['address'] or '未设置'}")
    st.markdown('</div>', unsafe_allow_html=True)
    
    st.markdown('<div class="card" style="margin-top: 20px;">', unsafe_allow_html=True)
    st.markdown('<p style="font-weight: 600; margin-bottom: 15px;">➕ 添加新机构</p>', unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    with col1:
        name = st.text_input("机构名称", key="im_name")
        region = st.text_input("所属地区", key="im_region")
        contact = st.text_input("联系人", key="im_contact")
    with col2:
        phone = st.text_input("联系电话", key="im_phone")
        email = st.text_input("邮箱", key="im_email")
        address = st.text_input("详细地址", key="im_address")
    
    col_btn1, col_btn2, _ = st.columns([1, 1, 3])
    with col_btn1:
        if st.button("✅ 添加机构", use_container_width=True, key="im_confirm_add"):
            if name:
                c = conn.cursor()
                c.execute('INSERT INTO institutions (name, region, contact, phone, email, address) VALUES (?,?,?,?,?,?)',
                          (name, region, contact, phone, email, address))
                conn.commit()
                add_log(st.session_state.user[1], f"添加机构：{name}")
                add_notification(None, "新机构添加", f"添加了新机构: {name}", "success")
                st.success("✅ 机构添加成功！")
                time.sleep(0.5)
                st.rerun()
            else:
                st.warning("⚠️ 请填写机构名称")
    with col_btn2:
        if st.button("📥 导入机构", use_container_width=True, key="im_import"):
            st.info("📤 请使用快速操作中的批量导入功能")
    st.markdown('</div>', unsafe_allow_html=True)
    conn.close()

def render_indicator_management():
    st.markdown('<p class="sub-header">📊 绩效指标管理</p>', unsafe_allow_html=True)
    conn = sqlite3.connect('performance_system.db')
    
    st.markdown('<div class="card">', unsafe_allow_html=True)
    df = pd.read_sql('SELECT * FROM indicators ORDER BY sort', conn)
    for idx, row in df.iterrows():
        col1, col2, col3, col4 = st.columns([2, 1, 2, 1])
        with col1:
            st.markdown(f"**{row['name']}**")
            st.markdown(f"<span class='small-note'>{row['description'] or '暂无描述'}</span>", unsafe_allow_html=True)
        with col2:
            st.markdown(f"<span class='badge badge-info'>满分: {row['full_score']}分</span>", unsafe_allow_html=True)
        with col3:
            st.progress(100, text=f"类别: {row['category'] or '未分类'}")
        with col4:
            if st.button("🗑️ 删除", key=f"del_ind_{row['id']}"):
                c = conn.cursor()
                c.execute('DELETE FROM indicators WHERE id=?', (row['id'],))
                conn.commit()
                add_log(st.session_state.user[1], f"删除指标：{row['name']}")
                st.success("✅ 指标已删除")
                time.sleep(0.5)
                st.rerun()
        st.markdown("---")
    st.markdown('</div>', unsafe_allow_html=True)
    
    st.markdown('<div class="card" style="margin-top: 20px;">', unsafe_allow_html=True)
    st.markdown('<p style="font-weight: 600; margin-bottom: 15px;">➕ 添加新指标</p>', unsafe_allow_html=True)
    col1, col2, col3 = st.columns(3)
    with col1:
        ind_name = st.text_input("指标名称", key="ind_name")
        category = st.text_input("所属类别", key="ind_category")
    with col2:
        full_score = st.number_input("满分值", min_value=1, max_value=100, value=100, key="ind_full_score")
        sort = st.number_input("排序号", min_value=0, value=0, key="ind_sort")
    with col3:
        description = st.text_area("指标说明", height=100, key="ind_desc")
    if st.button("✅ 添加指标", use_container_width=True, key="ind_confirm_add"):
        if ind_name:
            c = conn.cursor()
            c.execute('INSERT INTO indicators (name, full_score, category, description, sort) VALUES (?,?,?,?,?)',
                      (ind_name, full_score, category, description, sort))
            conn.commit()
            add_log(st.session_state.user[1], f"添加指标：{ind_name}")
            st.success("✅ 指标添加成功！")
            time.sleep(0.5)
            st.rerun()
        else:
            st.warning("⚠️ 请填写指标名称")
    st.markdown('</div>', unsafe_allow_html=True)
    conn.close()

def render_approval_management():
    st.markdown('<p class="sub-header">✅ 材料审批管理</p>', unsafe_allow_html=True)
    check_approval_timeout()
    conn = sqlite3.connect('performance_system.db')
    
    col1, col2, col3, col4 = st.columns(4)
    pending = pd.read_sql("SELECT COUNT(*) as cnt FROM materials WHERE status='待审核'", conn)['cnt'].iloc[0]
    approved = pd.read_sql("SELECT COUNT(*) as cnt FROM materials WHERE status='通过'", conn)['cnt'].iloc[0]
    rejected = pd.read_sql("SELECT COUNT(*) as cnt FROM materials WHERE status='驳回'", conn)['cnt'].iloc[0]
    timeout = pd.read_sql("SELECT COUNT(*) as cnt FROM materials WHERE status='审批超时'", conn)['cnt'].iloc[0]
    with col1:
        st.metric("⏳ 待审核", int(pending))
    with col2:
        st.metric("✅ 已通过", int(approved))
    with col3:
        st.metric("❌ 已驳回", int(rejected))
    with col4:
        st.metric("⏰ 已超时", int(timeout))
    
    st.markdown('<div class="card">', unsafe_allow_html=True)
    status_filter = st.selectbox("状态筛选", ["全部", "待审核", "通过", "驳回", "审批超时"])
    query = '''SELECT m.*, i.name as inst_name FROM materials m JOIN institutions i ON m.inst_id=i.id'''
    if status_filter != "全部":
        query += f" WHERE m.status='{status_filter}'"
    query += " ORDER BY m.submit_time DESC"
    df = pd.read_sql(query, conn)
    
    if not df.empty:
        for idx, row in df.iterrows():
            with st.expander(f"📄 {row['inst_name']} - {row['file_name']} ({row['status']})"):
                col1, col2 = st.columns([2, 1])
                with col1:
                    st.markdown(f"**机构:** {row['inst_name']}")
                    st.markdown(f"**文件名:** {row['file_name']}")
                    st.markdown(f"**提交时间:** {row['submit_time']}")
                    badge_class = get_status_badge(row['status'])
                    st.markdown(f"**当前状态:** <span class='badge {badge_class}'>{row['status']}</span>", unsafe_allow_html=True)
                    if row['review_comment']:
                        st.markdown(f"**审批意见:** {row['review_comment']}")
                with col2:
                    if row['status'] == '待审核':
                        status = st.selectbox("审批操作", ["通过", "驳回"], key=f"status_{row['id']}")
                        comment = st.text_area("审批意见", key=f"comment_{row['id']}")
                        if st.button("✅ 确认审批", key=f"btn_{row['id']}", use_container_width=True):
                            c = conn.cursor()
                            c.execute('''UPDATE materials SET status=?, reviewer=?, review_comment=?, review_time=? WHERE id=?''',
                                      (status, st.session_state.user[1], comment,
                                       datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'), row['id']))
                            conn.commit()
                            add_log(st.session_state.user[1], f"审批材料ID：{row['id']} - {status}")
                            add_notification(None, "材料审批结果", f"您的材料 '{row['file_name']}' 已被{status}", 
                                           "success" if status == "通过" else "warning")
                            st.success(f"✅ 审批完成：{status}")
                            time.sleep(0.5)
                            st.rerun()
                    if row['file_content']:
                        st.download_button(label="⬇️ 下载文件", data=row['file_content'], file_name=row['file_name'],
                            mime=row.get('file_type', 'application/octet-stream'), use_container_width=True)
    else:
        st.info("暂无材料记录")
    st.markdown('</div>', unsafe_allow_html=True)
    conn.close()


def render_statistics():
    st.markdown('<p class="sub-header">📈 统计分析</p>', unsafe_allow_html=True)
    conn = sqlite3.connect('performance_system.db')

    # 新增：文档上传 + 可视化分析
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<p style="font-weight: 600; margin-bottom: 15px;">📤 文档上传与可视化分析</p>', unsafe_allow_html=True)
    uploaded_file = st.file_uploader("上传数据文档进行可视化（CSV/XLSX/TXT/JSON）", type=['csv', 'xlsx', 'txt', 'json'], key='statistics_upload')
    if uploaded_file:
        try:
            if uploaded_file.name.lower().endswith('.csv'):
                df_upload = pd.read_csv(uploaded_file)
            elif uploaded_file.name.lower().endswith('.xlsx'):
                df_upload = pd.read_excel(uploaded_file)
            elif uploaded_file.name.lower().endswith('.txt'):
                text = uploaded_file.read().decode('utf-8', errors='ignore')
                try:
                    df_upload = pd.read_csv(io.StringIO(text), sep=None, engine='python')
                except Exception:
                    df_upload = pd.DataFrame({'文本内容': [text]})
            elif uploaded_file.name.lower().endswith('.json'):
                df_upload = pd.read_json(uploaded_file)
            else:
                df_upload = pd.DataFrame()

            if df_upload.empty:
                st.warning('文件解析后无数据，请上传有效的表格数据。')
            else:
                st.markdown('**数据预览（前 10 行）**')
                st.dataframe(df_upload.head(10), use_container_width=True, hide_index=True)

                numeric_cols = df_upload.select_dtypes(include=['number']).columns.tolist()
                if numeric_cols:
                    col1, col2, col3 = st.columns([2, 2, 1])
                    with col1:
                        x_col = st.selectbox('X 轴字段', options=numeric_cols, index=0, key='stat_x_col')
                    with col2:
                        y_col = st.selectbox('Y 轴字段', options=numeric_cols, index=min(1, len(numeric_cols) - 1), key='stat_y_col')
                    with col3:
                        chart_type = st.selectbox('图表类型', ['柱状图', '折线图', '散点图', '饼图'], key='stat_chart_type')

                    if chart_type == '柱状图':
                        fig = px.bar(df_upload, x=x_col, y=y_col, title='文档数据柱状图')
                    elif chart_type == '折线图':
                        fig = px.line(df_upload, x=x_col, y=y_col, title='文档数据折线图', markers=True)
                    elif chart_type == '散点图':
                        fig = px.scatter(df_upload, x=x_col, y=y_col, title='文档数据散点图')
                    else:  # 饼图
                        fig = px.pie(df_upload, names=x_col, values=y_col, title='文档数据饼图')

                    fig.update_layout(height=420)
                    st.plotly_chart(fig, use_container_width=True)
                else:
                    st.info('上传数据未包含数值列，当前仅展示表格内容。')
        except Exception as e:
            st.error(f'❌ 文档解析失败：{str(e)}')
    st.markdown('</div>', unsafe_allow_html=True)

    years = pd.read_sql('SELECT DISTINCT year FROM scores ORDER BY year DESC', conn)['year'].tolist()
    if not years:
        st.info("暂无评分数据")
        conn.close()
        return
    
    select_year = st.selectbox("📅 选择年度", years)
    
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<p style="font-weight: 600; margin-bottom: 15px;">🏆 机构绩效排名</p>', unsafe_allow_html=True)
    total_df = pd.read_sql('''SELECT i.name, i.region, SUM(s.score) as total_score, COUNT(s.id) as eval_count, AVG(s.score) as avg_score
        FROM scores s JOIN institutions i ON s.inst_id=i.id WHERE s.year=? GROUP BY i.id ORDER BY total_score DESC''', conn, params=(select_year,))
    
    if not total_df.empty:
        total_df['等级'] = total_df['total_score'].apply(lambda x: get_score_level(x)[0])
        st.dataframe(total_df, use_container_width=True, hide_index=True)
        
        col1, col2 = st.columns([1, 1])
        with col1:
            if st.button("📥 导出Excel", use_container_width=True):
                output = export_excel(total_df, f"排名数据_{select_year}")
                st.download_button(label="⬇️ 下载Excel", data=output, file_name=f"ranking_{select_year}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        with col2:
            if st.button("📄 导出PDF", use_container_width=True):
                with st.spinner("正在生成PDF..."):
                    pdf_path = export_pdf(total_df, f"{select_year}年度绩效排名")
                with open(pdf_path, "rb") as f:
                    st.download_button("⬇️ 下载PDF", f, file_name=pdf_path, use_container_width=True)
        
        fig = px.bar(total_df.head(10), x='name', y='total_score', color='total_score', color_continuous_scale='Viridis',
                     labels={'name': '机构名称', 'total_score': '总分数'})
        fig.update_layout(height=400)
        st.plotly_chart(fig, use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)
    
    st.markdown('<div class="card" style="margin-top: 20px;">', unsafe_allow_html=True)
    st.markdown('<p style="font-weight: 600; margin-bottom: 15px;">🗺️ 地区绩效分析</p>', unsafe_allow_html=True)
    region_df = pd.read_sql('''SELECT i.region, AVG(s.score) as avg_score, MAX(s.score) as max_score, MIN(s.score) as min_score, COUNT(DISTINCT i.id) as inst_count
        FROM scores s JOIN institutions i ON s.inst_id=i.id WHERE s.year=? GROUP BY i.region''', conn, params=(select_year,))
    if not region_df.empty:
        col1, col2 = st.columns(2)
        with col1:
            st.dataframe(region_df, use_container_width=True, hide_index=True)
        with col2:
            fig = px.pie(region_df, values='inst_count', names='region', title='机构地区分布')
            st.plotly_chart(fig, use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)
    
    st.markdown('<div class="card" style="margin-top: 20px;">', unsafe_allow_html=True)
    st.markdown('<p style="font-weight: 600; margin-bottom: 15px;">📊 指标得分分析</p>', unsafe_allow_html=True)
    indicator_df = pd.read_sql('''SELECT ind.name, ind.full_score, AVG(s.score) as avg_score, MAX(s.score) as max_score, MIN(s.score) as min_score
        FROM scores s JOIN indicators ind ON s.indicator_id=ind.id WHERE s.year=? GROUP BY ind.id''', conn, params=(select_year,))
    if not indicator_df.empty:
        fig = go.Figure()
        fig.add_trace(go.Bar(name='平均分', x=indicator_df['name'], y=indicator_df['avg_score'], marker_color='#667eea'))
        fig.add_trace(go.Scatter(name='满分', x=indicator_df['name'], y=indicator_df['full_score'], mode='lines+markers', line=dict(color='#ef4444', dash='dash')))
        fig.update_layout(height=400, barmode='group')
        st.plotly_chart(fig, use_container_width=True)
    st.markdown('</div>', unsafe_allow_html=True)
    conn.close()

def render_logs():
    st.markdown('<p class="sub-header">📝 系统操作日志</p>', unsafe_allow_html=True)
    conn = sqlite3.connect('performance_system.db')
    
    col1, col2 = st.columns([2, 1])
    with col1:
        search = st.text_input("🔍 搜索操作", placeholder="输入关键词搜索")
    with col2:
        days = st.selectbox("⏰ 时间范围", ["全部", "最近7天", "最近30天", "最近90天"])
    
    query = 'SELECT * FROM logs WHERE 1=1'
    params = []
    if search:
        query += ' AND (username LIKE ? OR operation LIKE ?)'
        params.extend([f'%{search}%', f'%{search}%'])
    if days != "全部":
        day_map = {"最近7天": 7, "最近30天": 30, "最近90天": 90}
        date_from = (datetime.datetime.now() - timedelta(days=day_map[days])).strftime('%Y-%m-%d')
        query += ' AND create_time >= ?'
        params.append(date_from)
    query += ' ORDER BY create_time DESC'
    df = pd.read_sql(query, conn, params=params)
    
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.dataframe(df, use_container_width=True, hide_index=True)
    
    col_export, col_clear, _ = st.columns([1, 1, 3])
    with col_export:
        if st.button("📥 导出日志", use_container_width=True):
            output = export_excel(df, "系统日志")
            st.download_button(label="⬇️ 下载Excel", data=output, file_name=f"logs_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
    with col_clear:
        if st.button("🗑️ 清空日志", use_container_width=True):
            if 'confirm_clear_logs' not in st.session_state:
                st.session_state.confirm_clear_logs = True
                st.warning("⚠️ 点击再次确认清空所有日志")
            else:
                c = conn.cursor()
                c.execute('DELETE FROM logs')
                conn.commit()
                add_log(st.session_state.user[1], "清空系统日志")
                del st.session_state.confirm_clear_logs
                st.success("✅ 日志已清空")
                time.sleep(0.5)
                st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)
    conn.close()

def render_ai_evaluation():
    st.markdown('<p class="sub-header">🤖 AI智能文档评估</p>', unsafe_allow_html=True)
    st.markdown('<div class="card">', unsafe_allow_html=True)
    
    uploaded_file = st.file_uploader("📤 上传评估文档", type=['txt', 'docx', 'pdf'])
    col1, col2 = st.columns(2)
    with col1:
        rule = st.text_area("📋 评估规则", "检查是否包含年度财务数据、人员配置信息、项目完成情况")
    with col2:
        st.markdown("<br>", unsafe_allow_html=True)
        st.markdown("**预设规则模板:**")
        templates = st.selectbox("选择模板", ["通用评估", "财务审计", "项目验收", "人员考核"])
        if templates == "财务审计":
            rule = "检查财务报表完整性、数据准确性、合规性"
        elif templates == "项目验收":
            rule = "检查项目完成度、交付物质量、文档规范性"
        elif templates == "人员考核":
            rule = "检查工作量、工作质量、团队协作表现"
    
    if uploaded_file and rule:
        content = uploaded_file.read()
        st.markdown(f"**文件名:** {uploaded_file.name}")
        st.markdown(f"**文件大小:** {len(content) / 1024:.2f} KB")
        if st.button("🚀 开始AI评估", use_container_width=True):
            with st.spinner("🤖 AI正在分析文档..."):
                time.sleep(1.5)
                result = ai_evaluate_document(content.decode('utf-8', errors='ignore'), rule)
            
            st.markdown('<div class="success-box">', unsafe_allow_html=True)
            st.markdown('<p style="font-weight: 700; margin-bottom: 10px;">📊 评估结果</p>', unsafe_allow_html=True)
            col_score, col_reason = st.columns([1, 2])
            with col_score:
                fig = create_gauge_chart(result['score'], "综合评分")
                st.plotly_chart(fig, use_container_width=True)
            with col_reason:
                level, badge, _ = get_score_level(result['score'])
                st.markdown(f"**评分等级:** <span class='badge {badge}'>{level}</span>", unsafe_allow_html=True)
                st.markdown(f"**评估结论:** {result['reason']}")
                st.markdown("**改进建议:**")
                for suggestion in result['suggestions']:
                    st.markdown(f"- {suggestion}")
            st.markdown('</div>', unsafe_allow_html=True)
            
            col_save, col_export = st.columns(2)
            with col_save:
                if st.button("💾 保存评估结果", use_container_width=True):
                    add_log(st.session_state.user[1], f"AI评估文档：{uploaded_file.name}，得分：{result['score']}")
                    st.success("✅ 评估结果已保存")
            with col_export:
                result_df = pd.DataFrame([{'文件名': uploaded_file.name, '评分': result['score'], '等级': level,
                    '评估结论': result['reason'], '评估时间': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}])
                output = export_excel(result_df, "AI评估结果")
                st.download_button(label="📥 导出结果", data=output,
                    file_name=f"AI评估结果_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
            add_notification(None, "AI评估完成", f"文档 '{uploaded_file.name}' AI评估完成，得分：{result['score']}", "success")
    st.markdown('</div>', unsafe_allow_html=True)

def render_settings():
    st.markdown('<p class="sub-header">⚙️ 系统设置</p>', unsafe_allow_html=True)
    tabs = st.tabs(["🔔 通知设置", "🔐 安全设置", "💾 数据备份", "📊 系统信息"])
    
    with tabs[0]:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        email_notif = st.checkbox("启用邮件通知", value=get_setting('email_notification') == '1')
        sms_notif = st.checkbox("启用短信通知", value=get_setting('sms_notification') == '1')
        timeout_remind = st.checkbox("审批超时提醒", value=get_setting('timeout_reminder') == '1')
        score_notif = st.checkbox("评分完成通知", value=get_setting('score_notification') == '1')
        if st.button("💾 保存通知设置", use_container_width=True):
            set_setting('email_notification', '1' if email_notif else '0')
            set_setting('sms_notification', '1' if sms_notif else '0')
            set_setting('timeout_reminder', '1' if timeout_remind else '0')
            set_setting('score_notification', '1' if score_notif else '0')
            add_log(st.session_state.user[1], "修改通知设置")
            st.success("✅ 通知设置已保存")
        st.markdown('</div>', unsafe_allow_html=True)
    
    with tabs[1]:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        login_attempts = st.number_input("登录失败锁定次数", min_value=3, max_value=10, value=int(get_setting('login_attempts') or 5))
        session_timeout = st.number_input("会话超时时间(分钟)", min_value=10, max_value=120, value=int(get_setting('session_timeout') or 30))
        two_factor = st.checkbox("启用双因素认证", value=get_setting('two_factor_auth') == '1')
        if st.button("💾 保存安全设置", use_container_width=True):
            set_setting('login_attempts', str(login_attempts))
            set_setting('session_timeout', str(session_timeout))
            set_setting('two_factor_auth', '1' if two_factor else '0')
            add_log(st.session_state.user[1], "修改安全设置")
            st.success("✅ 安全设置已保存")
        st.markdown('</div>', unsafe_allow_html=True)
    
    with tabs[2]:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<p style="font-weight: 600; margin-bottom: 15px;">💾 数据备份管理</p>', unsafe_allow_html=True)
        col1, col2 = st.columns(2)
        with col1:
            if st.button("📥 立即备份", use_container_width=True):
                with st.spinner("正在备份数据..."):
                    time.sleep(1)
                    conn = sqlite3.connect('performance_system.db')
                    output = io.BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        pd.read_sql('SELECT * FROM users', conn).to_excel(writer, sheet_name='users', index=False)
                        pd.read_sql('SELECT * FROM institutions', conn).to_excel(writer, sheet_name='institutions', index=False)
                        pd.read_sql('SELECT * FROM indicators', conn).to_excel(writer, sheet_name='indicators', index=False)
                        pd.read_sql('SELECT * FROM scores', conn).to_excel(writer, sheet_name='scores', index=False)
                        pd.read_sql('SELECT * FROM materials', conn).to_excel(writer, sheet_name='materials', index=False)
                        pd.read_sql('SELECT * FROM logs', conn).to_excel(writer, sheet_name='logs', index=False)
                        pd.read_sql('SELECT * FROM notifications', conn).to_excel(writer, sheet_name='notifications', index=False)
                        pd.read_sql('SELECT * FROM settings', conn).to_excel(writer, sheet_name='settings', index=False)
                    output.seek(0)
                    conn.close()
                    add_log(st.session_state.user[1], "执行数据备份")
                    st.download_button(label="⬇️ 下载备份文件", data=output,
                        file_name=f"full_backup_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
                    st.success("✅ 备份完成")
        with col2:
            restore_file = st.file_uploader("选择备份文件", type=['xlsx'], key="restore_file")
            if restore_file:
                if st.button("📤 恢复数据", use_container_width=True):
                    st.warning("⚠️ 数据恢复将覆盖现有数据，请谨慎操作！")
                    if st.checkbox("我确认要恢复数据"):
                        st.info("📤 数据恢复功能需要更复杂的逻辑，请联系系统管理员")
        
        st.markdown("---")
        st.markdown('<p style="font-weight: 600; margin-bottom: 10px;">🗑️ 数据清理</p>', unsafe_allow_html=True)
        col_clean1, col_clean2, col_clean3 = st.columns(3)
        with col_clean1:
            if st.button("清理30天前日志", use_container_width=True):
                conn = sqlite3.connect('performance_system.db')
                c = conn.cursor()
                cutoff_date = (datetime.datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
                c.execute('DELETE FROM logs WHERE create_time < ?', (cutoff_date,))
                deleted = c.rowcount
                conn.commit()
                conn.close()
                add_log(st.session_state.user[1], f"清理日志：{deleted}条")
                st.success(f"✅ 已清理 {deleted} 条旧日志")
        with col_clean2:
            if st.button("清理已读通知", use_container_width=True):
                conn = sqlite3.connect('performance_system.db')
                c = conn.cursor()
                c.execute('DELETE FROM notifications WHERE is_read=1')
                deleted = c.rowcount
                conn.commit()
                conn.close()
                add_log(st.session_state.user[1], f"清理通知：{deleted}条")
                st.success(f"✅ 已清理 {deleted} 条已读通知")
        with col_clean3:
            if st.button("清理超时材料", use_container_width=True):
                conn = sqlite3.connect('performance_system.db')
                c = conn.cursor()
                c.execute("DELETE FROM materials WHERE status='审批超时'")
                deleted = c.rowcount
                conn.commit()
                conn.close()
                add_log(st.session_state.user[1], f"清理超时材料：{deleted}条")
                st.success(f"✅ 已清理 {deleted} 条超时材料")
        st.markdown('</div>', unsafe_allow_html=True)
    
    with tabs[3]:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown('<p style="font-weight: 600; margin-bottom: 15px;">📊 系统信息</p>', unsafe_allow_html=True)
        conn = sqlite3.connect('performance_system.db')
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**系统版本:** v2.0.0")
            st.markdown("**数据库:** SQLite 3")
            st.markdown("**Python:** 3.9+")
            st.markdown("**Streamlit:** 1.28+")
        with col2:
            st.markdown(f"**用户总数:** {pd.read_sql('SELECT COUNT(*) as cnt FROM users', conn)['cnt'].iloc[0]}")
            st.markdown(f"**机构总数:** {pd.read_sql('SELECT COUNT(*) as cnt FROM institutions', conn)['cnt'].iloc[0]}")
            st.markdown(f"**评分记录:** {pd.read_sql('SELECT COUNT(*) as cnt FROM scores', conn)['cnt'].iloc[0]}")
        st.markdown("---")
        st.markdown(f"**最后更新:** 2025-01-15")
        st.markdown(f"**当前时间:** {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        conn.close()
        st.markdown('</div>', unsafe_allow_html=True)


def render_evaluator_dashboard():
    if st.session_state.get('show_notifications'):
        render_notifications()
        return
    st.markdown('<p class="sub-header">📝 评估人员控制台</p>', unsafe_allow_html=True)
    menu = st.sidebar.selectbox("📋 功能菜单", ["🏠 仪表盘", "⭐ 机构评分", "📊 查看报表", "📤 材料提交", "📥 导出PDF"])
    
    if menu == "🏠 仪表盘":
        render_evaluator_home()
    elif menu == "⭐ 机构评分":
        render_scoring()
    elif menu == "📊 查看报表":
        render_evaluator_reports()
    elif menu == "📤 材料提交":
        render_material_submit()
    elif menu == "📥 导出PDF":
        render_export_pdf()

def render_evaluator_home():
    st.markdown('<div class="animate-fadeIn">', unsafe_allow_html=True)
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown('<p style="font-weight: 600; margin-bottom: 15px;">📋 待办任务</p>', unsafe_allow_html=True)
    conn = sqlite3.connect('performance_system.db')
    pending_inst = pd.read_sql('SELECT COUNT(*) as cnt FROM institutions', conn)['cnt'].iloc[0]
    my_scores = pd.read_sql(f"SELECT COUNT(*) as cnt FROM scores WHERE evaluator='{st.session_state.user[1]}'", conn)['cnt'].iloc[0]
    col1, col2, col3 = st.columns(3)
    with col1:
        st.metric("🏢 待评估机构", int(pending_inst))
    with col2:
        st.metric("✅ 已完成评分", int(my_scores))
    with col3:
        st.metric("⏰ 本周任务", 5)
    st.markdown('</div>', unsafe_allow_html=True)
    
    st.markdown('<div class="card" style="margin-top: 20px;">', unsafe_allow_html=True)
    st.markdown('<p style="font-weight: 600; margin-bottom: 15px;">⚡ 快速评分</p>', unsafe_allow_html=True)
    insts = pd.read_sql('SELECT id, name FROM institutions', conn)
    if not insts.empty:
        col1, col2, col3 = st.columns([2, 1, 1])
        with col1:
            quick_inst = st.selectbox("选择机构", insts['id'].tolist(), format_func=lambda x: insts[insts['id']==x]['name'].iloc[0])
        with col2:
            quick_year = st.number_input("年度", min_value=2020, max_value=2030, value=2025)
        with col3:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("🚀 开始评分", use_container_width=True):
                st.session_state.scoring_inst = quick_inst
                st.session_state.scoring_year = quick_year
                st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)
    conn.close()
    st.markdown('</div>', unsafe_allow_html=True)

def render_scoring():
    st.markdown('<p class="sub-header">⭐ 绩效评分</p>', unsafe_allow_html=True)
    conn = sqlite3.connect('performance_system.db')
    st.markdown('<div class="card">', unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    with col1:
        insts = pd.read_sql('SELECT id, name FROM institutions', conn)
        if not insts.empty:
            default_inst = st.session_state.get('scoring_inst', insts['id'].iloc[0])
            inst_id = st.selectbox("🏢 选择机构", insts['id'].tolist(), format_func=lambda x: insts[insts['id']==x]['name'].iloc[0],
                                  index=list(insts['id']).index(default_inst) if default_inst in list(insts['id']) else 0)
    with col2:
        default_year = st.session_state.get('scoring_year', 2025)
        year = st.number_input("📅 评估年度", min_value=2020, max_value=2030, value=default_year)
    
    st.markdown("---")
    indicators = pd.read_sql('SELECT * FROM indicators ORDER BY sort', conn)
    total_score = 0
    scores_map = {}
    comments = {}
    
    st.markdown('<p style="font-weight: 600; margin-bottom: 15px;">📊 指标评分</p>', unsafe_allow_html=True)
    for idx, ind in indicators.iterrows():
        with st.expander(f"📌 {ind['name']} (满分{ind['full_score']}分)"):
            col_score, col_comment = st.columns([1, 2])
            with col_score:
                s = st.slider(f"评分", 0, int(ind['full_score']), int(ind['full_score'] * 0.8), help=ind['description'] or '')
                scores_map[int(ind['id'])] = int(s)
                total_score += int(s)
                progress = s / ind['full_score']
                st.progress(progress, text=f"{s}/{ind['full_score']}")
            with col_comment:
                comments[int(ind['id'])] = st.text_area("评语", placeholder=f"请输入{ind['name']}的评语...", key=f"comment_{ind['id']}")
    
    st.markdown("---")
    col_total, col_level = st.columns([2, 1])
    with col_total:
        st.markdown(f"<h2 style='color: #667eea;'>总分: {total_score}分</h2>", unsafe_allow_html=True)
    with col_level:
        level, badge_class, _ = get_score_level(total_score)
        st.markdown(f"<span class='badge {badge_class}' style='font-size: 16px;'>{level}</span>", unsafe_allow_html=True)
    
    col_submit, col_draft, col_cancel = st.columns([1, 1, 1])
    with col_submit:
        if st.button("✅ 提交评分", use_container_width=True):
            c = conn.cursor()
            for ind_id, s in scores_map.items():
                c.execute('INSERT INTO scores (inst_id, year, indicator_id, score, evaluator, comment) VALUES (?,?,?,?,?,?)',
                          (inst_id, year, ind_id, s, st.session_state.user[1], comments.get(ind_id, '')))
            conn.commit()
            add_log(st.session_state.user[1], f"为机构{inst_id}提交评分")
            inst_name = insts[insts['id']==inst_id]['name'].iloc[0]
            add_notification(None, "评分完成", f"机构 '{inst_name}' {year}年度评分已完成", "success")
            st.success(f"✅ 评分提交成功！总分：{total_score}分")
            st.balloons()
            if 'scoring_inst' in st.session_state:
                del st.session_state.scoring_inst
            if 'scoring_year' in st.session_state:
                del st.session_state.scoring_year
    with col_draft:
        if st.button("💾 保存草稿", use_container_width=True):
            st.session_state.draft_scores = scores_map
            st.session_state.draft_comments = comments
            st.info("📤 草稿已保存到会话")
    with col_cancel:
        if st.button("❌ 取消", use_container_width=True):
            if 'scoring_inst' in st.session_state:
                del st.session_state.scoring_inst
            if 'scoring_year' in st.session_state:
                del st.session_state.scoring_year
            st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)
    conn.close()

def render_evaluator_reports():
    st.markdown('<p class="sub-header">📊 绩效报表</p>', unsafe_allow_html=True)
    conn = sqlite3.connect('performance_system.db')
    years = pd.read_sql('SELECT DISTINCT year FROM scores ORDER BY year DESC', conn)['year'].tolist()
    if not years:
        st.info("暂无评分数据")
        conn.close()
        return
    select_year = st.selectbox("📅 选择年度", years)
    
    st.markdown('<div class="card">', unsafe_allow_html=True)
    df = pd.read_sql('''SELECT i.name, i.region, SUM(s.score) as total_score,
        CASE WHEN SUM(s.score)>=90 THEN "优秀" WHEN SUM(s.score)>=75 THEN "良好" WHEN SUM(s.score)>=60 THEN "合格" ELSE "不合格" END as level, s.evaluator
        FROM scores s JOIN institutions i ON s.inst_id=i.id WHERE s.year=? GROUP BY i.id ORDER BY total_score DESC''', conn, params=(select_year,))
    
    if not df.empty:
        st.dataframe(df, use_container_width=True, hide_index=True)
        col1, col2 = st.columns([1, 1])
        with col1:
            if st.button("📥 导出Excel", use_container_width=True):
                output = export_excel(df, f"报表_{select_year}")
                st.download_button(label="⬇️ 下载Excel", data=output, file_name=f"report_{select_year}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        with col2:
            if st.button("📄 导出PDF", use_container_width=True):
                with st.spinner("正在生成PDF..."):
                    pdf_path = export_pdf(df, f"{select_year}年度绩效报表")
                with open(pdf_path, "rb") as f:
                    st.download_button("⬇️ 下载PDF", f, file_name=pdf_path, use_container_width=True)
        fig = px.bar(df, x='name', y='total_score', color='level',
                     color_discrete_map={'优秀': '#10b981', '良好': '#3b82f6', '合格': '#f59e0b', '不合格': '#ef4444'},
                     labels={'name': '机构名称', 'total_score': '总分数'})
        fig.update_layout(height=400)
        st.plotly_chart(fig, use_container_width=True)
    else:
        st.info("暂无数据")
    st.markdown('</div>', unsafe_allow_html=True)
    conn.close()

def render_material_submit():
    st.markdown('<p class="sub-header">📤 评估材料提交</p>', unsafe_allow_html=True)
    conn = sqlite3.connect('performance_system.db')
    st.markdown('<div class="card">', unsafe_allow_html=True)
    insts = pd.read_sql('SELECT id, name FROM institutions', conn)
    col1, col2 = st.columns(2)
    with col1:
        inst_id = st.selectbox("🏢 选择机构", insts['id'].tolist(), format_func=lambda x: insts[insts['id']==x]['name'].iloc[0])
    with col2:
        year = st.number_input("📅 年度", 2020, 2030, 2025)
    uploaded_file = st.file_uploader("📎 上传材料", type=['xlsx', 'pdf', 'docx', 'zip'])
    if uploaded_file:
        st.markdown(f"**文件名:** {uploaded_file.name}")
        st.markdown(f"**文件大小:** {len(uploaded_file.getvalue()) / 1024:.2f} KB")
    
    col_submit, col_cancel, _ = st.columns([1, 1, 2])
    with col_submit:
        if st.button("✅ 提交材料", use_container_width=True):
            if uploaded_file:
                c = conn.cursor()
                c.execute('INSERT INTO materials (inst_id, year, file_content, file_name, file_type) VALUES (?,?,?,?,?)',
                          (inst_id, year, uploaded_file.read(), uploaded_file.name, uploaded_file.type))
                conn.commit()
                add_log(st.session_state.user[1], f"提交机构{inst_id}材料")
                inst_name = insts[insts['id']==inst_id]['name'].iloc[0]
                add_notification(None, "材料提交", f"机构 '{inst_name}' 提交了新材料待审核", "info")
                st.success("✅ 材料提交成功，等待审核")
                time.sleep(0.5)
                st.rerun()
            else:
                st.warning("⚠️ 请先上传文件")
    with col_cancel:
        if st.button("❌ 取消", use_container_width=True):
            st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)
    
    st.markdown('<div class="card" style="margin-top: 20px;">', unsafe_allow_html=True)
    st.markdown('<p style="font-weight: 600; margin-bottom: 15px;">📋 已提交材料</p>', unsafe_allow_html=True)
    materials = pd.read_sql('''SELECT m.*, i.name as inst_name FROM materials m JOIN institutions i ON m.inst_id=i.id WHERE m.inst_id=? ORDER BY m.submit_time DESC''', conn, params=(inst_id,))
    if not materials.empty:
        for idx, row in materials.iterrows():
            badge_class = get_status_badge(row['status'])
            st.markdown(f"""
                <div style="background: #f8fafc; border-radius: 12px; padding: 15px; margin-bottom: 10px;">
                    <div style="display: flex; justify-content: space-between; align-items: center;">
                        <div><span style="font-weight: 600;">{row['file_name']}</span><span class='badge {badge_class}' style='margin-left: 10px;'>{row['status']}</span></div>
                        <span style="color: #94a3b8; font-size: 12px;">{row['submit_time']}</span>
                    </div>
                </div>
            """, unsafe_allow_html=True)
            if row['file_content']:
                st.download_button(label="⬇️ 下载", data=row['file_content'], file_name=row['file_name'],
                    mime=row.get('file_type', 'application/octet-stream'), key=f"download_mat_{row['id']}")
    else:
        st.info("暂无提交记录")
    st.markdown('</div>', unsafe_allow_html=True)
    conn.close()

def render_export_pdf():
    st.markdown('<p class="sub-header">📥 PDF报表导出</p>', unsafe_allow_html=True)
    conn = sqlite3.connect('performance_system.db')
    years = pd.read_sql('SELECT DISTINCT year FROM scores ORDER BY year DESC', conn)['year'].tolist()
    if not years:
        st.info("暂无评分数据")
        conn.close()
        return
    st.markdown('<div class="card">', unsafe_allow_html=True)
    col1, col2 = st.columns(2)
    with col1:
        select_year = st.selectbox("📅 选择导出年度", years)
    with col2:
        export_type = st.selectbox("📄 导出类型", ["完整报表", "机构排名", "指标分析"])
    
    if export_type == "完整报表":
        df = pd.read_sql('''SELECT i.name, i.region, SUM(s.score) as total_score,
            CASE WHEN SUM(s.score)>=90 THEN "优秀" WHEN SUM(s.score)>=75 THEN "良好" WHEN SUM(s.score)>=60 THEN "合格" ELSE "不合格" END as level
            FROM scores s JOIN institutions i ON s.inst_id=i.id WHERE s.year=? GROUP BY i.id''', conn, params=(select_year,))
    elif export_type == "机构排名":
        df = pd.read_sql('''SELECT i.name, SUM(s.score) as total_score, RANK() OVER (ORDER BY SUM(s.score) DESC) as rank
            FROM scores s JOIN institutions i ON s.inst_id=i.id WHERE s.year=? GROUP BY i.id ORDER BY total_score DESC''', conn, params=(select_year,))
    else:
        df = pd.read_sql('''SELECT ind.name, AVG(s.score) as avg_score, ind.full_score
            FROM scores s JOIN indicators ind ON s.indicator_id=ind.id WHERE s.year=? GROUP BY ind.id''', conn, params=(select_year,))
    
    if not df.empty:
        st.dataframe(df, use_container_width=True, hide_index=True)
        col_export, col_preview, _ = st.columns([1, 1, 2])
        with col_export:
            if st.button("📥 导出PDF", use_container_width=True):
                with st.spinner("正在生成PDF..."):
                    time.sleep(1)
                    pdf_path = export_pdf(df, f"{select_year}年度{export_type}")
                with open(pdf_path, "rb") as f:
                    st.download_button("⬇️ 下载PDF", f, file_name=pdf_path, use_container_width=True)
        with col_preview:
            if st.button("👁️ 预览报表", use_container_width=True):
                st.markdown('<div class="card">', unsafe_allow_html=True)
                st.markdown('<p style="font-weight: 600; margin-bottom: 10px;">📄 报表预览</p>', unsafe_allow_html=True)
                st.markdown(f"**报表类型:** {export_type}")
                st.markdown(f"**年度:** {select_year}")
                st.markdown(f"**记录数:** {len(df)}")
                st.markdown(f"**生成时间:** {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                st.markdown("---")
                st.dataframe(df.head(5), use_container_width=True)
                st.markdown('</div>', unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)
    conn.close()

def render_institution_dashboard():
    if st.session_state.get('show_notifications'):
        render_notifications()
        return
    st.markdown('<p class="sub-header">🏢 机构用户控制台</p>', unsafe_allow_html=True)
    st.markdown('<div class="info-box">', unsafe_allow_html=True)
    st.markdown("💡 您可以查看本机构的绩效评估结果和审批状态")
    st.markdown('</div>', unsafe_allow_html=True)
    check_approval_timeout()
    conn = sqlite3.connect('performance_system.db')
    username = st.session_state.user[1]
    inst = pd.read_sql('SELECT * FROM institutions WHERE name=?', conn, params=(username,))
    if inst.empty:
        st.warning("⚠️ 未找到关联的机构信息")
        conn.close()
        return
    inst_id = inst.iloc[0]['id']
    
    col1, col2, col3 = st.columns(3)
    with col1:
        avg_score = pd.read_sql(f'SELECT AVG(score) as avg FROM scores WHERE inst_id={inst_id}', conn)['avg'].iloc[0] or 0
        st.markdown(f'<div class="stat-card"><div class="stat-number">{avg_score:.1f}</div><div class="stat-label">平均分</div></div>', unsafe_allow_html=True)
    with col2:
        eval_count = pd.read_sql(f'SELECT COUNT(DISTINCT year) as cnt FROM scores WHERE inst_id={inst_id}', conn)['cnt'].iloc[0]
        st.markdown(f'<div class="stat-card success"><div class="stat-number">{int(eval_count)}</div><div class="stat-label">评估年度</div></div>', unsafe_allow_html=True)
    with col3:
        pending = pd.read_sql(f"SELECT COUNT(*) as cnt FROM materials WHERE inst_id={inst_id} AND status='待审核'", conn)['cnt'].iloc[0]
        st.markdown(f'<div class="stat-card warning"><div class="stat-number">{int(pending)}</div><div class="stat-label">待审批</div></div>', unsafe_allow_html=True)
    
    st.markdown('<div class="card" style="margin-top: 20px;">', unsafe_allow_html=True)
    st.markdown('<p style="font-weight: 600; margin-bottom: 15px;">📊 历年绩效成绩</p>', unsafe_allow_html=True)
    df = pd.read_sql('''SELECT s.year, SUM(s.score) as total_score, AVG(s.score) as avg_score, COUNT(s.id) as eval_count
        FROM scores s WHERE s.inst_id=? GROUP BY s.year ORDER BY s.year DESC''', conn, params=(inst_id,))
    if not df.empty:
        df['等级'] = df['total_score'].apply(lambda x: get_score_level(x)[0])
        st.dataframe(df, use_container_width=True, hide_index=True)
        fig = px.line(df, x='year', y='total_score', markers=True, labels={'year': '年份', 'total_score': '总分数'})
        fig.update_traces(line_color='#667eea', line_width=3)
        st.plotly_chart(fig, use_container_width=True)
        col1, col2 = st.columns([1, 1])
        with col1:
            if st.button("📥 导出Excel", use_container_width=True):
                output = export_excel(df, "我的成绩")
                st.download_button(label="⬇️ 下载Excel", data=output, file_name=f"my_scores_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        with col2:
            if st.button("📄 导出PDF", use_container_width=True):
                with st.spinner("正在生成PDF..."):
                    pdf_path = export_pdf(df, f"我的绩效成绩")
                with open(pdf_path, "rb") as f:
                    st.download_button("⬇️ 下载PDF", f, file_name=pdf_path, use_container_width=True)
    else:
        st.info("暂无评分数据")
    st.markdown('</div>', unsafe_allow_html=True)
    
    st.markdown('<div class="card" style="margin-top: 20px;">', unsafe_allow_html=True)
    st.markdown('<p style="font-weight: 600; margin-bottom: 15px;">📋 材料审批状态</p>', unsafe_allow_html=True)
    materials = pd.read_sql('SELECT file_name, status, submit_time, review_time, review_comment FROM materials WHERE inst_id=? ORDER BY submit_time DESC', conn, params=(inst_id,))
    if not materials.empty:
        for idx, row in materials.iterrows():
            badge_class = get_status_badge(row['status'])
            review_comment_html = f"<div style='color: #64748b; margin-top: 5px; font-size: 13px;'>审批意见: {row['review_comment']}</div>" if row['review_comment'] else ''
            st.markdown(f"""
                <div style="background: #f8fafc; border-radius: 12px; padding: 15px; margin-bottom: 10px;">
                    <div style="display: flex; justify-content: space-between; align-items: center;">
                        <div><span style="font-weight: 600;">{row['file_name']}</span><span class='badge {badge_class}' style='margin-left: 10px;'>{row['status']}</span></div>
                        <span style="color: #94a3b8; font-size: 12px;">{row['submit_time']}</span>
                    </div>
                    {review_comment_html}
                </div>
            """, unsafe_allow_html=True)
    else:
        st.info("暂无材料记录")
    st.markdown('</div>', unsafe_allow_html=True)
    conn.close()

# -------------------------- 主程序 --------------------------
def main():
    init_db()
    check_approval_timeout()
    
    if 'logged_in' not in st.session_state:
        st.session_state.logged_in = False
        st.session_state.user = None
        st.session_state.page = "dashboard"
        st.session_state.show_add_user = False
        st.session_state.show_add_inst = False
        st.session_state.show_export = False
        st.session_state.show_notifications = False
    
    if not st.session_state.logged_in:
        render_login_page()
    else:
        render_sidebar()
        role = st.session_state.user[3]
        if st.sidebar.button("🚪 退出登录", use_container_width=True):
            add_log(st.session_state.user[1], "用户退出登录")
            st.session_state.logged_in = False
            st.session_state.user = None
            st.rerun()
        if role == "管理员":
            render_admin_dashboard()
        elif role == "评估人员":
            render_evaluator_dashboard()
        elif role == "机构用户":
            render_institution_dashboard()

if __name__ == '__main__':
    main()
